import argparse
import re
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.core.database import connect_database
from app.models import Branch, Company, Employee, User


PLACEHOLDER_PHONE = "0000000000"
DEFAULT_LATITUDE = 26.9124
DEFAULT_LONGITUDE = 75.7873


def clean(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text in {"-", "###########"}:
        return ""
    return text


def parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean(value)
    if not text:
        return None
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def split_name(name: str) -> tuple[str, str]:
    parts = clean(name).split()
    if not parts:
        return "Employee", "Staff"
    if len(parts) == 1:
        return parts[0], "Staff"
    return parts[0], " ".join(parts[1:])


def safe_branch_code(domain: str) -> str:
    code = re.sub(r"[^A-Za-z0-9]+", "-", domain.upper()).strip("-")
    return code or "BRANCH"


def valid_email(email: str, fallback_code: str) -> str:
    email = clean(email).lower()
    if re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", email):
        return email
    return f"{fallback_code.lower().replace('/', '-')}@employee.local"


def get_company_context():
    company_admin = User.objects(email="companyadmin@example.com").first()
    if company_admin and company_admin.company_id:
        company = Company.objects(id=company_admin.company_id).first()
        if company:
            return company.tenant_id, str(company.id)
    company = Company.objects(status="active").first() or Company.objects.first()
    if not company:
        raise RuntimeError("No company found. Create company/admin first.")
    return company.tenant_id, str(company.id)


def main():
    parser = argparse.ArgumentParser(description="Import employees from Employee Details Excel sheet.")
    parser.add_argument("xlsx_path")
    parser.add_argument("--apply", action="store_true", help="Create branches and employees.")
    args = parser.parse_args()

    connect_database()
    tenant_id, company_id = get_company_context()
    workbook = load_workbook(args.xlsx_path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [clean(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    column = {name: index for index, name in enumerate(headers)}

    rows = []
    branch_counts = defaultdict(int)
    for raw in sheet.iter_rows(min_row=2, values_only=True):
        domain = clean(raw[column["Domain"]]).upper()
        name = clean(raw[column["Name"]])
        if not domain or not name:
            continue
        branch_counts[domain] += 1
        count = branch_counts[domain]
        employee_code = f"A2G/{safe_branch_code(domain)}/{count:04d}"
        first_name, last_name = split_name(name)
        rows.append(
            {
                "branch_domain": domain,
                "employee_code": employee_code,
                "first_name": first_name,
                "last_name": last_name,
                "father_name": clean(raw[column["Father Name"]]),
                "mother_name": clean(raw[column["Mother Name"]]),
                "email": valid_email(raw[column["Email Address"]], employee_code),
                "phone": PLACEHOLDER_PHONE,
                "emergency_contact_number": PLACEHOLDER_PHONE,
                "designation": clean(raw[column["Designation"]]) or "Staff",
                "joining_date": parse_date(raw[column["Joining Date"]]),
                "date_of_birth": parse_date(raw[column["DOB"]]),
                "current_address": clean(raw[column["Address"]]),
                "permanent_address": clean(raw[column["Address"]]),
                "pan_number": clean(raw[column["PAN Number"]]),
                "note": f"Imported from employee sheet. Location: {clean(raw[column['Location']]) or 'N/A'}",
            }
        )

    print(f"Rows ready: {len(rows)}")
    print("Branches:")
    for branch, count in sorted(branch_counts.items()):
        exists = Branch.objects(tenant_id=tenant_id, company_id=company_id, branch_code=safe_branch_code(branch)).first()
        print(f"  - {branch}: {count} employees, {'exists' if exists else 'will create'}")

    if not args.apply:
        print("\nDry run only. Run with --apply to import.")
        return

    created_branches = 0
    created_employees = 0
    updated_employees = 0
    used_emails = set()
    for item in rows:
        branch_code = safe_branch_code(item.pop("branch_domain"))
        branch = Branch.objects(tenant_id=tenant_id, company_id=company_id, branch_code=branch_code).first()
        if not branch:
            branch = Branch(
                tenant_id=tenant_id,
                company_id=company_id,
                branch_name=branch_code,
                branch_code=branch_code,
                address=f"{branch_code} Branch",
                latitude=DEFAULT_LATITUDE,
                longitude=DEFAULT_LONGITUDE,
                allowed_radius=100,
                kiosk_pin="1234",
            ).save()
            created_branches += 1

        email = item["email"]
        if email in used_emails or Employee.objects(email=email).first():
            email = f"{item['employee_code'].lower().replace('/', '-')}@employee.local"
        used_emails.add(email)

        payload = {
            **item,
            "email": email,
            "office_email": email,
            "tenant_id": tenant_id,
            "company_id": company_id,
            "branch_id": branch,
            "status": "active",
        }
        existing = Employee.objects(tenant_id=tenant_id, company_id=company_id, employee_code=item["employee_code"]).first()
        if existing:
            for key, value in payload.items():
                setattr(existing, key, value)
            existing.save()
            updated_employees += 1
        else:
            Employee(**payload).save()
            created_employees += 1

    print("\nImport complete.")
    print(f"Created branches: {created_branches}")
    print(f"Created employees: {created_employees}")
    print(f"Updated employees: {updated_employees}")


if __name__ == "__main__":
    main()
