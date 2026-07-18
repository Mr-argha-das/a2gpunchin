import argparse
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.core.database import connect_database
from app.models import Employee, Shift, User


def clean(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text in {"-", "###########"}:
        return ""
    return text


def phone(value) -> str:
    digits = re.sub(r"\D+", "", clean(value))
    return digits or "0000000000"


def company_context():
    user = User.objects(email="companyadmin@example.com").first() or User.objects(is_super_admin=True).first()
    if not user:
        raise RuntimeError("Admin user not found.")
    return user.tenant_id, user.company_id


def main():
    parser = argparse.ArgumentParser(description="Assign Office Main Shift and update employee phone numbers from sheet.")
    parser.add_argument("xlsx_path")
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    connect_database()
    tenant_id, company_id = company_context()
    shift = Shift.objects(tenant_id=tenant_id, company_id=company_id, shift_name="Office Main Shift").first()
    if not shift:
        shift = Shift(
            tenant_id=tenant_id,
            company_id=company_id,
            shift_name="Office Main Shift",
            start_time="09:00",
            end_time="18:00",
            grace_time=10,
            late_after=15,
            half_day_after=240,
            early_logout_before=30,
        )
        if args.apply:
            shift.save()

    workbook = load_workbook(args.xlsx_path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [clean(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    column = {name: index for index, name in enumerate(headers)}
    branch_counts = {}
    updates = []

    for raw in sheet.iter_rows(min_row=2, values_only=True):
        domain = clean(raw[column["Domain"]]).upper()
        name = clean(raw[column["Name"]])
        if not domain or not name:
            continue
        branch_counts[domain] = branch_counts.get(domain, 0) + 1
        branch_code = re.sub(r"[^A-Za-z0-9]+", "-", domain.upper()).strip("-") or "BRANCH"
        employee_code = f"A2G/{branch_code}/{branch_counts[domain]:04d}"
        updates.append(
            {
                "employee_code": employee_code,
                "phone": phone(raw[column["Phone Number"]]),
                "emergency_contact_number": phone(raw[column["Emergency Contact Number"]]),
            }
        )

    found = 0
    missing = []
    for update in updates:
        employee = Employee.objects(tenant_id=tenant_id, company_id=company_id, employee_code=update["employee_code"]).first()
        if not employee:
            missing.append(update["employee_code"])
            continue
        found += 1
        if args.apply:
            employee.phone = update["phone"]
            employee.emergency_contact_number = update["emergency_contact_number"]
            employee.shift_id = shift
            employee.save()

    print(f"Shift: Office Main Shift ({'existing' if shift.id else 'will create'})")
    print(f"Employees matched: {found}")
    print(f"Employees missing: {len(missing)}")
    if missing[:10]:
        print("First missing:", ", ".join(missing[:10]))
    print("Applied" if args.apply else "Dry run only. Run with --apply to update.")


if __name__ == "__main__":
    main()
