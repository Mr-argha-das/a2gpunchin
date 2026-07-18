import sys
from pathlib import Path

from openpyxl import load_workbook


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Usage: python scripts/inspect_employee_sheet.py <xlsx_path>")
    path = Path(sys.argv[1])
    workbook = load_workbook(path, read_only=True, data_only=True)
    print(f"Workbook: {path}")
    print(f"Sheets: {', '.join(workbook.sheetnames)}")
    for sheet in workbook.worksheets:
        print(f"\nSheet: {sheet.title} rows={sheet.max_row} cols={sheet.max_column}")
        rows = list(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 8), values_only=True))
        for row in rows:
            print([value for value in row])


if __name__ == "__main__":
    main()
